from __future__ import annotations

import csv
from pathlib import Path

from kv.converters.base import BaseConverter, ConvertResult


class ExcelConverter(BaseConverter):
    source_type = "excel"
    extensions = {".xlsx", ".xlsm", ".xltx", ".csv", ".tsv"}

    def convert(self, path: Path) -> ConvertResult:
        if path.suffix.lower() in {".csv", ".tsv"}:
            tables = self._csv_to_markdown(path)
            kind = "CSV → 마크다운 표"
        else:
            tables = self._sheets_to_markdown(path)
            kind = "엑셀 → 마크다운 표"
        body = f"""# {path.stem}

## 원본
- 파일: `{path.name}`
- 유형: {kind}

{tables}
"""
        return ConvertResult(
            title=path.stem,
            body=body,
            source_type=self.source_type,
            extra_meta={"sheet_count": tables.count("## 시트:")},
        )

    def _csv_to_markdown(self, path: Path) -> str:
        try:
            delim = "\t" if path.suffix.lower() == ".tsv" else ","
            with open(path, encoding="utf-8-sig", newline="") as f:
                rows = [tuple(r) for r in csv.reader(f, delimiter=delim)]
            if not rows:
                return "(빈 CSV 파일)"
            return self._rows_to_md_table(rows)
        except Exception as e:
            return f"(CSV 변환 오류: {e})"

    def _sheets_to_markdown(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            parts: list[str] = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                parts.append(f"## 시트: {name}\n")
                parts.append(self._rows_to_md_table(rows))
                parts.append("")
            wb.close()
            return "\n".join(parts) if parts else "(빈 엑셀 파일)"
        except Exception as e:
            return f"(엑셀 변환 오류: {e})"

    @staticmethod
    def _rows_to_md_table(rows: list[tuple]) -> str:
        clean = [[("" if c is None else str(c)).replace("|", "\\|") for c in row] for row in rows]
        if not clean:
            return ""
        width = max(len(r) for r in clean)
        padded = [r + [""] * (width - len(r)) for r in clean]
        header = padded[0]
        sep = ["---"] * width
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(sep) + " |",
        ]
        for row in padded[1:]:
            lines.append("| " + " | ".join(row) + " |")
        return "\n".join(lines)
