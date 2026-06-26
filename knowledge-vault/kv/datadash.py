"""업로드한 표(엑셀·CSV)를 자동 분석 → '자료 대시보드'.

컬럼을 숫자/분류로 자동 판별해:
- 숫자 컬럼: 합계·평균·최대·최소
- 분류 컬럼: 값별 개수(상위)
파일·시트별로 요약을 만든다. (도메인 무관 — 아무 표나)
"""

from __future__ import annotations

import csv as _csv
import re
from pathlib import Path

from kv.config import INBOX

_TABLE_EXT = {".xlsx", ".xlsm", ".xls", ".csv", ".tsv"}


def _num(v):
    """문자열을 숫자로 (콤마·공백 제거). 실패 시 None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _read_rows(path: Path) -> list[tuple[str, list[list]]]:
    """(시트명, 행들[리스트of리스트]) 목록. 첫 행은 헤더."""
    ext = path.suffix.lower()
    out: list[tuple[str, list[list]]] = []
    if ext in {".csv", ".tsv"}:
        delim = "\t" if ext == ".tsv" else ","
        try:
            with open(path, encoding="utf-8-sig", newline="") as f:
                rows = [r for r in _csv.reader(f, delimiter=delim)]
            out.append((path.stem, rows))
        except Exception:
            pass
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            for name in wb.sheetnames:
                ws = wb[name]
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                if rows:
                    out.append((name, rows))
            wb.close()
        except Exception:
            pass
    return out


def _summarize_sheet(rows: list[list]) -> dict | None:
    # 헤더 = 첫 비어있지 않은 행
    rows = [r for r in rows if any((c is not None and str(c).strip()) for c in r)]
    if len(rows) < 2:
        return None
    header = [str(c).strip() if c is not None else f"열{i+1}" for i, c in enumerate(rows[0])]
    data = rows[1:]
    ncol = len(header)
    cols: list[dict] = []
    for ci in range(ncol):
        vals = [r[ci] if ci < len(r) else None for r in data]
        nums = [n for n in (_num(v) for v in vals) if n is not None]
        nonempty = [str(v).strip() for v in vals if v is not None and str(v).strip()]
        if not nonempty:
            continue
        name = header[ci] or f"열{ci+1}"
        # 숫자 컬럼: 값의 70% 이상이 숫자
        if nonempty and len(nums) >= max(1, int(len(nonempty) * 0.7)):
            cols.append({
                "name": name, "type": "number",
                "sum": round(sum(nums), 2), "avg": round(sum(nums) / len(nums), 2),
                "min": min(nums), "max": max(nums), "count": len(nums),
            })
        else:
            counts: dict[str, int] = {}
            for v in nonempty:
                counts[v] = counts.get(v, 0) + 1
            top = sorted(counts.items(), key=lambda x: -x[1])[:6]
            cols.append({
                "name": name, "type": "category",
                "unique": len(counts), "top": top,
            })
    return {"rows": len(data), "columns": cols}


def summarize() -> list[dict]:
    """inbox 의 모든 표 파일 요약."""
    out: list[dict] = []
    if not INBOX.exists():
        return out
    for path in sorted(INBOX.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in _TABLE_EXT:
            continue
        if path.name.startswith("~") or path.name.startswith("."):
            continue
        sheets = []
        for sheet_name, rows in _read_rows(path):
            s = _summarize_sheet(rows)
            if s:
                s["sheet"] = sheet_name
                sheets.append(s)
        if sheets:
            out.append({"file": path.name, "sheets": sheets})
    return out
